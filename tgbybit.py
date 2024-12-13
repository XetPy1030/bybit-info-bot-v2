import requests
import telebot
from telebot import types
from datetime import datetime, timedelta
from time import sleep
import threading
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import logging
import hashlib
import hmac
import time
import sys
import json
import os
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator

logging.basicConfig(level=logging.ERROR)
plt.switch_backend('Agg')

CONFIG_FILE = 'config.json'


def load_config():
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_config(config):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4, ensure_ascii=False)

config = load_config()

USE_API = config.get('USE_API', False)
TOKEN = config.get('TOKEN', '')
API_KEY = config.get('API_KEY', '')
API_SECRET = config.get('API_SECRET', '')
cookies = config.get('cookies', '')
admins = config.get('admins', [])
db_update_interval = config.get('db_update_interval', 30)
balance_send_interval = config.get('balance_send_interval', 30)
chat_id = config.get('chat_id', '')

REQUEST_TIMEOUT = 60
MAX_RETRIES = 5
EXCEL_FILE = 'balance_data.xlsx'
RECV_WINDOW = 10000
WAITING_FOR_RENEW = False
BYBIT_DOMAINS = [
    "https://api.bybit.com"
]

bot = telebot.TeleBot(TOKEN)

BALANCE_URL = 'https://api2.bybit.com/v3/private/cht/asset-common/total-balance?quoteCoin=USDT&balanceType=1'
BOT_LIST_URL = 'https://api2.bybit.com/s1/bot/tradingbot/v1/list-all-bots'

BASE_URL = None



keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
keyboard.add(types.KeyboardButton('/balance'), types.KeyboardButton('/graph'))

last_balance = None

db_update_thread = None
balance_send_thread = None
stop_threads = False

def setup_excel():
    try:
        workbook = load_workbook(EXCEL_FILE)
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Дата', 'Баланс USDT', 'Баланс RUB', 'Изменение (%)'])
        workbook.save(EXCEL_FILE)
    return workbook, worksheet

workbook, worksheet = setup_excel()

def expire_mode_notify():
    global WAITING_FOR_RENEW
    WAITING_FOR_RENEW = True
    for admin_id in admins:
        try:
            bot.send_message(admin_id, "Срок действия данных истёк или возникла ошибка соединения. Обновите данные и перезапустите бота через панель админа.")
        except:
            pass

def retry_request(url, method='GET', headers=None, params=None, cookies_arg=None, timeout=REQUEST_TIMEOUT):
    attempts = 0
    while attempts < MAX_RETRIES:
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers, params=params, cookies=cookies_arg, timeout=timeout)
            else:
                response = requests.post(url, headers=headers, data=params, cookies=cookies_arg, timeout=timeout)

            response.raise_for_status()
            return response
        except requests.RequestException as e:
            logging.error(f"Ошибка запроса: {e}")
            attempts += 1
            sleep(2 ** attempts)

    expire_mode_notify()
    return None

def get_usdt_to_rub():
    if WAITING_FOR_RENEW:
        return None
    response = retry_request('https://api.coingecko.com/api/v3/simple/price?ids=tether&vs_currencies=rub')
    if response and not WAITING_FOR_RENEW:
        data = response.json()
        return float(data['tether']['rub'])
    return None

def initialize_api():
    global BASE_URL
    for domain in BYBIT_DOMAINS:
        try:
            resp = requests.get(f"{domain}/v5/public/time", timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                if data.get('retCode') == 0:
                    BASE_URL = domain
                    return True
        except Exception as e:
            logging.error(f"Не удалось подключиться к {domain}: {e}")
    logging.error("Не удалось найти доступный домен Bybit API.")
    expire_mode_notify()
    return False

def get_server_time():
    if BASE_URL:
        response = retry_request(f'{BASE_URL}/v5/public/time')
        if response and not WAITING_FOR_RENEW:
            data = response.json()
            if data.get('retCode') == 0:
                return int(data['result']['time'])
    return int(time.time() * 1000)

def generate_signature(secret, timestamp, api_key, recv_window, params_str):
    pre_sign_str = str(timestamp) + api_key + str(recv_window) + params_str
    return hmac.new(secret.encode('utf-8'), pre_sign_str.encode('utf-8'), hashlib.sha256).hexdigest()


def fetch_bot_list_data():
    if WAITING_FOR_RENEW:
        return []
    response = retry_request(BOT_LIST_URL, method='POST', cookies_arg={'secure-token': cookies})
    if response and not WAITING_FOR_RENEW:
        data = response.json()
        if data.get("ret_code") == 0:
            bots = data.get("result", {}).get("bots", [])
            return bots
    return []


def fetch_balance_cookies(add_to_db=True):
    global last_balance
    if WAITING_FOR_RENEW:
        return "Бот в режиме ожидания обновления данных."

    response = retry_request(
        BALANCE_URL,
        cookies_arg={'secure-token': cookies})
    if response and not WAITING_FOR_RENEW:
        data = response.json()
        if 'result' in data and 'totalBalanceItems' in data['result']:
            for item in data['result']['totalBalanceItems']:
                if item['accountType'] == 'ACCOUNT_TYPE_BOT':
                    current_balance = float(item['originBalance'])
                    usdt_to_rub = get_usdt_to_rub()
                    if usdt_to_rub:
                        rub_balance = current_balance * usdt_to_rub
                    else:
                        rub_balance = "Ошибка курса"

                    now = datetime.now()

                    rows = list(worksheet.iter_rows(values_only=True))[1:]
                    closest_balance_24h_ago = None
                    closest_time_diff = float('inf')
                    twenty_four_hours_ago_ts = time.time() - 24*3600

                    for row in rows:
                        timestamp_row = datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
                        time_diff = abs(timestamp_row.timestamp() - twenty_four_hours_ago_ts)
                        if time_diff < closest_time_diff:
                            closest_time_diff = time_diff
                            closest_balance_24h_ago = float(row[1])

                    change_percent = 0
                    if closest_balance_24h_ago is not None:
                        change_percent = ((current_balance - closest_balance_24h_ago) / closest_balance_24h_ago) * 100

                    if add_to_db:
                        last_balance = current_balance
                        worksheet.append([now.strftime('%Y-%m-%d %H:%M:%S'), current_balance, rub_balance, change_percent])
                        workbook.save(EXCEL_FILE)

                    change_direction = "📈" if change_percent >= 0 else "📉"
                    change_color = "🟢" if change_percent >= 0 else "🔴"
                    change_percent_str = f"{change_color} {'+' if change_percent >= 0 else ''}{change_percent:.2f}%"

                    return (f"📅 Дата: {now.strftime('%Y-%m-%d %H:%M:%S')}\n"
                            f"💰 Баланс: {current_balance} USDT = {rub_balance:.2f} RUB\n"
                            f"{change_direction} Изменение за 24ч: {change_percent_str}")
        else:
            expire_mode_notify()
            return "Срок действия cookies истёк. Бот в ожидании."
    return "Ошибка соединения или данные недоступны"


def fetch_balance_api(add_to_db=True):
    global last_balance
    if WAITING_FOR_RENEW:
        return "Бот в режиме ожидания обновления данных."
    if not BASE_URL and not initialize_api():
        return "Ошибка инициализации API. Бот в ожидании."

    url = f'{BASE_URL}/v5/account/wallet-balance'
    account_type = 'UNIFIED'
    timestamp = get_server_time()
    query_params = {'accountType': account_type}

    params_str = "accountType=" + account_type
    sign = generate_signature(API_SECRET, timestamp, API_KEY, RECV_WINDOW, params_str)

    headers = {
        "X-BAPI-API-KEY": API_KEY,
        "X-BAPI-TIMESTAMP": str(timestamp),
        "X-BAPI-SIGN": sign,
        "X-BAPI-RECV-WINDOW": str(RECV_WINDOW)
    }

    response = retry_request(url, headers=headers, params=query_params)
    if response and not WAITING_FOR_RENEW:
        data = response.json()
        ret_code = data.get('retCode', None)
        if ret_code == 0:
            result = data.get('result', {})
            lists = result.get('list', [])
            current_balance = None

            for acc in lists:
                if acc.get('accountType') == account_type:
                    for coin_info in acc.get('coin', []):
                        if coin_info.get('coin') == 'USDT':
                            current_balance = float(coin_info.get('equity', 0))
                            break

            if current_balance is None:
                expire_mode_notify()
                return "Ошибка: Баланс USDT не найден. Бот в ожидании."

            usdt_to_rub = get_usdt_to_rub()
            if usdt_to_rub:
                rub_balance = current_balance * usdt_to_rub
            else:
                rub_balance = "Ошибка курса"

            now = datetime.now()
            rows = list(worksheet.iter_rows(values_only=True))[1:]
            closest_balance_24h_ago = None
            closest_time_diff = float('inf')
            twenty_four_hours_ago_ts = time.time() - 24*3600
            for row in rows:
                timestamp_row = datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
                time_diff = abs(timestamp_row.timestamp() - twenty_four_hours_ago_ts)
                if time_diff < closest_time_diff:
                    closest_time_diff = time_diff
                    closest_balance_24h_ago = float(row[1])

            change_percent = 0
            if closest_balance_24h_ago is not None:
                change_percent = ((current_balance - closest_balance_24h_ago) / closest_balance_24h_ago) * 100

            if add_to_db:
                last_balance = current_balance
                worksheet.append([now.strftime('%Y-%m-%d %H:%M:%S'), current_balance, rub_balance, change_percent])
                workbook.save(EXCEL_FILE)

            change_direction = "📈" if change_percent >= 0 else "📉"
            change_color = "🟢" if change_percent >= 0 else "🔴"
            change_percent_str = f"{change_color} {'+' if change_percent >= 0 else ''}{change_percent:.2f}%"

            return (f"📅 Дата: {now.strftime('%Y-%m-%d %H:%M:%S')}\n"
                    f"💰 Баланс: {current_balance} USDT = {rub_balance:.2f} RUB\n"
                    f"{change_direction} Изменение за 24ч: {change_percent_str}")
        else:
            ret_msg = data.get('retMsg', 'Неизвестная ошибка')
            expire_mode_notify()
            return f"Ошибка API: {ret_msg}. Бот в ожидании."
    expire_mode_notify()
    return "Ошибка соединения или данные недоступны. Бот в ожидании."


def fetch_balance(add_to_db=True):
    if USE_API:
        balance_info = fetch_balance_api(add_to_db=add_to_db)
    else:
        balance_info = fetch_balance_cookies(add_to_db=add_to_db)
    if isinstance(balance_info, str) and "Баланс:" in balance_info:
        rows = list(worksheet.iter_rows(values_only=True))[1:]
        now_ts = time.time()
        count_24h = sum(
            1 for r in rows if (now_ts - datetime.strptime(r[0], '%Y-%m-%d %H:%M:%S').timestamp()) <= 24 * 3600)
        if len(rows) > 1:
            last_balance_val = rows[-1][1]
            prev_balance_val = rows[-2][1]
            diff = last_balance_val - prev_balance_val
            diff_str = f"Изменение с последнего замера: {'+' if diff >= 0 else ''}{diff:.2f} USDT"
        else:
            diff_str = "Недостаточно данных для расчёта изменения с последнего замера."

        balance_info += f"\n📊 Записей за последние 24ч: {count_24h}\n{diff_str}"
    return balance_info


@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    bot.send_message(message.chat.id, "Бот запущен", reply_markup=keyboard)


@bot.message_handler(commands=['balance'])
def balance_cmd(message):
    try:
        balance_info = fetch_balance(add_to_db=False)
        bot.send_message(message.chat.id, balance_info)
    except Exception as e:
        logging.error(f"Ошибка при отправке баланса: {e}")
        bot.send_message(message.chat.id, "Произошла ошибка при отправке баланса")


@bot.message_handler(commands=['graph'])
def send_graph(message):
    try:
        rows = list(worksheet.iter_rows(values_only=True))[1:]
        if len(rows) < 2:
            bot.send_message(message.chat.id, "Недостаточно данных для построения графика")
            return
        daily_balances = {}
        current_day_balances = []
        now_date = datetime.now().date()

        for row in rows:
            timestamp_str = row[0]
            try:
                timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                continue
            date = timestamp.date()
            balance_usdt = row[1]
            if isinstance(balance_usdt, (int, float)):
                if date not in daily_balances:
                    daily_balances[date] = []
                daily_balances[date].append(balance_usdt)
                if date == now_date:
                    current_day_balances.append((timestamp, balance_usdt))
        average_daily_balances = []
        max_daily_balances = []
        min_daily_balances = []
        for date, balances in daily_balances.items():
            avg = sum(balances) / len(balances)
            maximum = max(balances)
            minimum = min(balances)
            average_daily_balances.append((date, avg))
            max_daily_balances.append((date, maximum))
            min_daily_balances.append((date, minimum))
        average_daily_balances.sort(key=lambda x: x[0])
        if len(average_daily_balances) > 30:
            average_daily_balances_last_month = average_daily_balances[-30:]
            max_daily_balances_last_month = max_daily_balances[-30:]
            min_daily_balances_last_month = min_daily_balances[-30:]
        else:
            average_daily_balances_last_month = average_daily_balances
            max_daily_balances_last_month = max_daily_balances
            min_daily_balances_last_month = min_daily_balances

        dates_last_month = [item[0] for item in average_daily_balances_last_month]
        average_balances_usdt_last_month = [item[1] for item in average_daily_balances_last_month]
        max_balances_last_month = [item[1] for item in max_daily_balances_last_month]
        min_balances_last_month = [item[1] for item in min_daily_balances_last_month]
        one_year_ago = now_date - timedelta(days=365)
        monthly_balances = {}
        for date, balances in daily_balances.items():
            if date >= one_year_ago:
                month = date.replace(day=1)
                if month not in monthly_balances:
                    monthly_balances[month] = []
                monthly_balances[month].extend(balances)

        average_monthly_balances_year = []
        max_monthly_balances_year = []
        min_monthly_balances_year = []
        for month, balances in monthly_balances.items():
            avg = sum(balances) / len(balances)
            maximum = max(balances)
            minimum = min(balances)
            average_monthly_balances_year.append((month, avg))
            max_monthly_balances_year.append((month, maximum))
            min_monthly_balances_year.append((month, minimum))
        average_monthly_balances_year.sort(key=lambda x: x[0])

        dates_year = [item[0] for item in average_monthly_balances_year]
        average_balances_usdt_year = [item[1] for item in average_monthly_balances_year]
        max_balances_year = [item[1] for item in max_monthly_balances_year]
        min_balances_year = [item[1] for item in min_monthly_balances_year]
        if len(current_day_balances) < 2:
            bot.send_message(message.chat.id, "Недостаточно данных за текущий день для построения детального графика")
            return

        current_day_balances.sort(key=lambda x: x[0])
        times = [item[0] for item in current_day_balances]
        balances_usdt = [item[1] for item in current_day_balances]
        bots_data = fetch_bot_list_data()
        fig = plt.figure(figsize=(20, 12), tight_layout=True)
        gs = fig.add_gridspec(3, 2, width_ratios=[4, 1], height_ratios=[2, 2, 2], wspace=0.2, hspace=0.3)
        ax1 = fig.add_subplot(gs[0, 0])
        ax2 = fig.add_subplot(gs[1, 0])
        ax3 = fig.add_subplot(gs[2, 0])
        ax4 = fig.add_subplot(gs[:, 1])
        ax4.axis('off')

        y_locator = MaxNLocator(nbins=10)
        ax1.yaxis.set_major_locator(y_locator)
        ax2.yaxis.set_major_locator(y_locator)
        ax3.yaxis.set_major_locator(y_locator)

        ax1.plot(times, balances_usdt, marker='o', linestyle='-', color='tab:red', label='Баланс (текущий день)')
        ax1.set_ylabel('Баланс (USDT)')
        ax1.set_title('Баланс за текущий день', fontsize=14)
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        ax1.tick_params(axis='x', rotation=45)
        ax1.grid(True, which='both', linestyle='--', linewidth=0.5)
        for i, txt in enumerate(balances_usdt):
            ax1.annotate(f'{txt:.2f}', (times[i], balances_usdt[i]),
                        textcoords="offset points", xytext=(0, 10), ha='center', fontsize=8)
        ax1.legend()

        ax2.plot(dates_last_month, average_balances_usdt_last_month, marker='o', linestyle='-', color='tab:blue', label='Средний баланс (30 дней)')
        ax2.set_ylabel('Средний баланс (USDT)')
        ax2.set_title('Средний баланс за последние 30 дней', fontsize=14)
        ax2.xaxis.set_major_formatter(mdates.DateFormatter('%d'))
        ax2.tick_params(axis='x', rotation=45)
        ax2.grid(True, which='both', linestyle='--', linewidth=0.5)
        for i in range(len(dates_last_month)):
            ax2.annotate(
                f'{max_balances_last_month[i]:.2f}\n{average_balances_usdt_last_month[i]:.2f}\n{min_balances_last_month[i]:.2f}',
                (dates_last_month[i], average_balances_usdt_last_month[i]),
                textcoords="offset points", xytext=(0, 10), ha='center', fontsize=7
            )
        ax2.legend()

        if average_monthly_balances_year:
            ax3.plot(dates_year, average_balances_usdt_year, marker='o', linestyle='-', color='tab:green', label='Средний баланс (год)')
            ax3.set_ylabel('Средний баланс (USDT)')
            ax3.set_title('Средний баланс за последний год', fontsize=14)
            ax3.xaxis.set_major_formatter(mdates.DateFormatter('%b'))
            ax3.tick_params(axis='x', rotation=45)
            ax3.grid(True, which='both', linestyle='--', linewidth=0.5)
            for i in range(len(dates_year)):
                ax3.annotate(
                    f'{max_balances_year[i]:.2f}\n{average_balances_usdt_year[i]:.2f}\n{min_balances_year[i]:.2f}',
                    (dates_year[i], average_balances_usdt_year[i]),
                    textcoords="offset points", xytext=(0, 10), ha='center', fontsize=7
                )
            ax3.legend()
        else:
            ax3.text(0.5, 0.5, 'Нет данных за последний год', horizontalalignment='center',
                     verticalalignment='center', transform=ax3.transAxes)

        bot_info_lines = []
        if bots_data:
            bot_info_lines.append("Информация о ботах:")
            for idx, trading_bot in enumerate(bots_data):
                b_type = trading_bot.get('type', 'N/A')
                symbol = 'N/A'
                invested = 'N/A'
                pnl = 'N/A'
                pnl_per = '0.00%'
                price_range = 'N/A'
                price_drop = 'N/A'
                cell_num = 'N/A'
                add_pos_per = 'N/A'

                if b_type == 'GRID_FUTURES' and trading_bot.get('future_grid'):
                    fg = trading_bot['future_grid']
                    symbol = fg.get('symbol', 'N/A')
                    invested = fg.get('total_investment', 'N/A')
                    pnl = fg.get('pnl', 'N/A')
                    try:
                        pnl_value = float(fg.get('pnl_per', '0'))
                        pnl_per = f"{pnl_value * 100:.2f}%"
                    except:
                        pnl_per = "0.00%"
                    price_range = f"{fg.get('min_price', 'N/A')}/{fg.get('max_price', 'N/A')}"
                    price_drop = fg.get('liq_price', 'N/A')
                    cell_num = fg.get('cell_num', 'N/A')
                    add_pos_per = fg.get('leverage', 'N/A')

                elif b_type == 'MART_FUTURES' and trading_bot.get('fmart'):
                    fmtr = trading_bot['fmart']
                    symbol = fmtr.get('symbol', 'N/A')
                    invested = fmtr.get('total_margin', 'N/A')
                    pnl = fmtr.get('total_profit', 'N/A')
                    try:
                        pnl_value = float(fmtr.get('total_profit_per', '0'))
                        pnl_per = f"{pnl_value * 100:.2f}%"
                    except:
                        pnl_per = "0.00%"
                    price_range = 'N/A'
                    price_drop = 'N/A'
                    cell_num = 'N/A'
                    add_pos_per = fmtr.get('add_pos_per', 'N/A')

                elif b_type == 'GRID_SPOT' and trading_bot.get('grid', {}).get('info'):
                    gr = trading_bot['grid']['info']
                    profit = trading_bot['grid']['profit']
                    symbol = gr.get('symbol', 'N/A')
                    invested = gr.get('total_investment', 'N/A')
                    pnl = profit.get('total_profit', 'N/A')
                    try:
                        pnl_value = float(profit.get('total_apr', '0'))
                        pnl_per = f"{pnl_value * 100:.2f}%"
                    except:
                        pnl_per = "0.00%"
                    price_range = f"{gr.get('min_price', 'N/A')}/{gr.get('max_price', 'N/A')}"
                    price_drop = gr.get('liq_price', 'N/A')
                    cell_num = gr.get('cell_number', 'N/A')
                    add_pos_per = 'N/A'

                bot_text = (
                    f"*{symbol}*\n" 
                    f"Инвестировано (USDT): {invested}\n"
                    f"Общий P&L (USDT): {pnl}\n"
                    f"Ценовой диапазон (USDT): {price_range}\n"
                    f"Снижение цены: {price_drop}\n"
                    f"Кол-во сеток: {cell_num}\n"
                    f"Множитель позиции: {add_pos_per}\n"
                    f"% PnL: *{pnl_per}*"
                )

                bot_info_lines.append(bot_text)

        else:
            bot_info_lines.append("Нет данных о ботах.")

        bot_info_text = "\n\n".join(bot_info_lines)

        ax4.text(0.02, 0.98, bot_info_text, ha='left', va='top', wrap=True, fontsize=10,
                 bbox=dict(facecolor='lightgray', edgecolor='gray', boxstyle='round,pad=0.5'),
                 transform=ax4.transAxes, color='black', fontweight='normal')

        graph_filename = 'graph.png'
        plt.savefig(graph_filename, dpi=300)
        plt.close()

        with open(graph_filename, 'rb') as photo:
            bot.send_photo(message.chat.id, photo, parse_mode='Markdown')
    except Exception as e:
        print(f"Ошибка генерации графика: {e}")
        bot.send_message(message.chat.id, "Произошла ошибка при генерации графика.")


def wait_until_next_interval(minutes):
    now = datetime.now()
    minute = (now.minute // minutes + 1) * minutes
    hour = now.hour
    if minute >= 60:
        minute = 0
        hour = (hour + 1) % 24
    target = datetime(now.year, now.month, now.day, hour, minute, 0)
    delta = (target - now).total_seconds()
    if delta < 0:
        target += timedelta(days=1)
        delta = (target - now).total_seconds()
    sleep(delta)

def db_update_loop():
    while not stop_threads:
        if not WAITING_FOR_RENEW:
            fetch_balance()
        wait_until_next_interval(db_update_interval)

def balance_send_loop():
    while not stop_threads:
        if not WAITING_FOR_RENEW:
            balance_info = fetch_balance(add_to_db=False)
            if isinstance(balance_info, str) and chat_id:
                try:
                    bot.send_message(chat_id, balance_info)
                except:
                    pass
        wait_until_next_interval(balance_send_interval)


threads_started = False

def start_threads():
    global db_update_thread, balance_send_thread, stop_threads, threads_started
    if threads_started:
        return
    stop_threads = False
    db_update_thread = threading.Thread(target=db_update_loop, daemon=True)
    balance_send_thread = threading.Thread(target=balance_send_loop, daemon=True)
    db_update_thread.start()
    balance_send_thread.start()
    threads_started = True

def stop_all_threads():
    global stop_threads, threads_started
    stop_threads = True
    threads_started = False

def is_admin(user_id):
    return user_id in admins

@bot.message_handler(commands=['admin'])
def admin_panel(message):
    if message.chat.type != 'private':
        return
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "У вас нет прав доступа.")
        return

    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Изменить TOKEN", callback_data="change_token"))
    markup.add(types.InlineKeyboardButton("Изменить API_KEY", callback_data="change_api_key"))
    markup.add(types.InlineKeyboardButton("Изменить API_SECRET", callback_data="change_api_secret"))
    markup.add(types.InlineKeyboardButton("Изменить cookies", callback_data="change_cookies"))
    markup.add(types.InlineKeyboardButton(f"Переключить USE_API (сейчас {USE_API})", callback_data="toggle_use_api"))
    markup.add(types.InlineKeyboardButton("Скачать базу данных", callback_data="download_db"))
    markup.add(types.InlineKeyboardButton("Изменить интервал обновления БД", callback_data="change_db_interval"))
    markup.add(types.InlineKeyboardButton("Изменить интервал отправки баланса", callback_data="change_balance_interval"))
    markup.add(types.InlineKeyboardButton("Добавить админа", callback_data="add_admin"))
    markup.add(types.InlineKeyboardButton("Удалить админа", callback_data="remove_admin"))
    markup.add(types.InlineKeyboardButton("Показать текущие настройки", callback_data="show_config"))
    markup.add(types.InlineKeyboardButton("Перезапустить бота", callback_data="reload_bot"))
    markup.add(types.InlineKeyboardButton("Снять режим ожидания", callback_data="resume_bot" if WAITING_FOR_RENEW else "no_wait_mode"))

    bot.send_message(message.chat.id, "Панель админа:", reply_markup=markup)


pending_actions = {}

@bot.callback_query_handler(func=lambda call: True)
def callback_admin(call):
    bot.answer_callback_query(call.id)
    user_id = call.from_user.id
    chat_type = call.message.chat.type if call.message else None
    if chat_type != 'private':
        return
    if not is_admin(user_id):
        return

    if call.data in ["change_token", "change_api_key", "change_api_secret", "change_cookies",
                     "change_db_interval", "change_balance_interval", "add_admin", "remove_admin"]:
        pending_actions[user_id] = (call.data,)
        field_name = {
            "change_token": "TOKEN",
            "change_api_key": "API_KEY",
            "change_api_secret": "API_SECRET",
            "change_cookies": "cookies",
            "change_db_interval": "интервал обновления БД (мин)",
            "change_balance_interval": "интервал отправки баланса (мин)",
            "add_admin": "ID нового админа",
            "remove_admin": "ID админа для удаления"
        }[call.data]
        bot.send_message(user_id, f"Отправьте новое значение для: {field_name}")

    elif call.data == "toggle_use_api":
        global USE_API
        USE_API = not USE_API
        config['USE_API'] = USE_API
        save_config(config)
        bot.send_message(user_id, f"USE_API переключен. Сейчас USE_API={USE_API}")
    elif call.data == "download_db":
        if os.path.exists(EXCEL_FILE):
            with open(EXCEL_FILE, 'rb') as f:
                bot.send_document(user_id, f)
        else:
            bot.send_message(user_id, "Файл базы данных не найден.")
    elif call.data == "show_config":
        conf_text = (
            f"Текущие настройки:\n\n"
            f"USE_API: {config.get('USE_API', False)}\n"
            f"TOKEN: {config.get('TOKEN', '')}\n"
            f"API_KEY: {config.get('API_KEY', '')}\n"
            f"API_SECRET: {config.get('API_SECRET', '')}\n"
            f"cookies: {config.get('cookies', '')}\n"
            f"admins: {config.get('admins', [])}\n"
            f"db_update_interval: {config.get('db_update_interval', 30)} минут\n"
            f"balance_send_interval: {config.get('balance_send_interval', 30)} минут\n"
            f"chat_id: {config.get('chat_id', '')}"
        )
        bot.send_message(user_id, conf_text)
    elif call.data == "reload_bot":
        reload_config()
        bot.send_message(user_id, "Конфиг перезагружен, бот работает с новыми параметрами.")
    elif call.data == "resume_bot":
        global WAITING_FOR_RENEW
        WAITING_FOR_RENEW = False
        bot.send_message(user_id, "Режим ожидания снят, бот продолжит работу.")
    elif call.data == "no_wait_mode":
        bot.send_message(user_id, "Бот не в режиме ожидания.")


def reload_config():
    global config, USE_API, TOKEN, API_KEY, API_SECRET, cookies, admins, db_update_interval, balance_send_interval, chat_id
    global stop_threads, db_update_thread, balance_send_thread

    config = load_config()
    USE_API = config.get('USE_API', False)
    TOKEN = config.get('TOKEN', '')
    API_KEY = config.get('API_KEY', '')
    API_SECRET = config.get('API_SECRET', '')
    cookies = config.get('cookies', '')
    admins = config.get('admins', [])
    db_update_interval = config.get('db_update_interval', 30)
    balance_send_interval = config.get('balance_send_interval', 30)
    chat_id = config.get('chat_id', '')

    stop_all_threads()
    sleep(1)
    start_threads()


@bot.message_handler(func=lambda message: message.from_user.id in pending_actions)
def admin_input_handler(message):
    user_id = message.from_user.id
    action = pending_actions[user_id][0]

    try:
        if action == "change_token":
            config['TOKEN'] = message.text.strip()
            bot.send_message(user_id, "TOKEN обновлён.")
        elif action == "change_api_key":
            config['API_KEY'] = message.text.strip()
            bot.send_message(user_id, "API_KEY обновлён.")
        elif action == "change_api_secret":
            config['API_SECRET'] = message.text.strip()
            bot.send_message(user_id, "API_SECRET обновлён.")
        elif action == "change_cookies":
            config['cookies'] = message.text.strip()
            bot.send_message(user_id, "cookies обновлены.")
        elif action == "change_db_interval":
            interval = int(message.text.strip())
            config['db_update_interval'] = interval
            bot.send_message(user_id, f"Интервал обновления БД теперь {interval} минут.")
        elif action == "change_balance_interval":
            interval = int(message.text.strip())
            config['balance_send_interval'] = interval
            bot.send_message(user_id, f"Интервал отправки баланса теперь {interval} минут.")
        elif action == "add_admin":
            new_admin = int(message.text.strip())
            if new_admin not in config['admins']:
                config['admins'].append(new_admin)
                bot.send_message(user_id, f"Админ {new_admin} добавлен.")
            else:
                bot.send_message(user_id, f"{new_admin} уже админ.")
        elif action == "remove_admin":
            remove_id = int(message.text.strip())
            if remove_id in config['admins']:
                config['admins'].remove(remove_id)
                bot.send_message(user_id, f"Админ {remove_id} удалён.")
            else:
                bot.send_message(user_id, f"{remove_id} не найден в списке админов.")

        save_config(config)

    except ValueError:
        bot.send_message(user_id, "Некорректный ввод.")

    del pending_actions[user_id]


if __name__ == '__main__':
    if USE_API:
        initialize_api()

    start_threads()
    bot.polling(non_stop=True)
